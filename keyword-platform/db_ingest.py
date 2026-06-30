#!/usr/bin/env python3
"""
将重点词 xlsx 表格数据导入云数据库（Postgres / Supabase / Neon）。

幂等：同一 (日期, 关键词) 重复导入会覆盖，不会产生重复行。

用法：
    # 连接串放在环境变量或同目录 .env 文件里（见 .env.example）
    python3 db_ingest.py                                 # 导入所有 *.xlsx，平台默认"小程序"
    python3 db_ingest.py 文件.xlsx                        # 只导入指定文件（小程序）
    python3 db_ingest.py --platform APP app数据.xlsx      # 指定平台为 APP

平台说明：
    历史数据均为"小程序"，默认即可；APP 数据导入时加 --platform APP。
    若某个 xlsx 自带"平台"列，则按该列逐行取值，--platform 仅作兜底默认。

依赖：pip install openpyxl psycopg2-binary
"""

import os
import sys
import glob
import openpyxl
import psycopg2
from psycopg2.extras import execute_values

DIR = os.path.dirname(os.path.abspath(__file__))


def load_dotenv():
    """读取同目录 .env（KEY=VALUE），不覆盖已存在的环境变量。"""
    path = os.path.join(DIR, '.env')
    if not os.path.exists(path):
        return
    with open(path, encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith('#') or '=' not in line:
                continue
            k, v = line.split('=', 1)
            k, v = k.strip(), v.strip().strip('"').strip("'")
            os.environ.setdefault(k, v)


def get_conn():
    load_dotenv()
    url = os.environ.get('DATABASE_URL')
    if not url:
        sys.exit("缺少连接串：请设置环境变量 DATABASE_URL，或在本目录创建 .env（参考 .env.example）")
    return psycopg2.connect(url)


# 表头中文名 -> 解析方式
def find_cols(header):
    """返回各字段的列索引（按表头名定位，容忍列顺序变化）。"""
    idx = {}
    for i, h in enumerate(header):
        if not h:
            continue
        h = str(h)
        if h.startswith('日期'):            idx['date'] = i
        elif 'Division' in h:               idx['division'] = i
        elif h == '关键词' or h.startswith('关键词') and 'Division' not in h:
            idx.setdefault('keyword', i)
        elif '搜索人数' in h:               idx['search_users'] = i
        elif '搜索次数' in h:               idx['search_count'] = i
        elif h.upper().startswith('GMV'):   idx['gmv'] = i
        elif '加购' in h:                   idx['add_cart_rate'] = i
        elif '支付率' in h:                 idx['pay_rate'] = i
        elif '无结果' in h:                 idx['no_result_rate'] = i
        elif '平台' in h:                   idx['platform'] = i
    missing = {'date', 'keyword'} - set(idx)
    if missing:
        sys.exit(f"表头缺少必要列：{missing}（实际表头：{header}）")
    return idx


def num(v):
    if v is None or v == '':
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def pick_sheet(wb):
    """优先选名字含'关键词表现'/'关键词'的 sheet，否则用 active。"""
    for name in wb.sheetnames:
        if '关键词表现' in name:
            return wb[name]
    for name in wb.sheetnames:
        if '关键词' in name:
            return wb[name]
    return wb.active


def parse_xlsx(path, default_platform='小程序'):
    """返回 (rows, stats)。
    平台取值：xlsx 自带"平台"列则逐行取，否则用 default_platform。"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = pick_sheet(wb)
    it = ws.iter_rows(values_only=True)
    header = next(it)
    c = find_cols(header)
    fname = os.path.basename(path)

    rows, dates, kws, plats = [], set(), set(), set()
    for r in it:
        if c['date'] >= len(r) or not r[c['date']] or not r[c['keyword']]:
            continue
        date = str(r[c['date']])[:10]
        kw = str(r[c['keyword']]).strip()
        def cell(key, cast=None):
            i = c.get(key)
            if i is None or i >= len(r):
                return None
            v = r[i]
            if cast and v is not None:
                try: return cast(v)
                except (TypeError, ValueError): return None
            return v
        pi = c.get('platform')
        plat = (str(r[pi]).strip() if pi is not None and pi < len(r) and r[pi] else None) or default_platform
        rows.append((
            date, kw, plat,
            (str(r[c['division']]).strip() if c.get('division') is not None and r[c['division']] else None),
            cell('search_users', lambda x: int(round(float(x)))),
            cell('search_count', lambda x: int(round(float(x)))),
            num(cell('gmv')),
            num(cell('add_cart_rate')),
            num(cell('pay_rate')),
            num(cell('no_result_rate')),
            fname,
        ))
        dates.add(date); kws.add(kw); plats.add(plat)
    wb.close()
    stats = dict(file=fname, n=len(rows),
                 dmin=min(dates) if dates else None,
                 dmax=max(dates) if dates else None,
                 kw=len(kws),
                 platform='/'.join(sorted(plats)) if plats else default_platform)
    return rows, stats


UPSERT = """
INSERT INTO keyword_daily
    (date, keyword, platform, division, search_users, search_count, gmv,
     add_cart_rate, pay_rate, no_result_rate, source_file)
VALUES %s
ON CONFLICT (date, keyword, platform) DO UPDATE SET
    division       = EXCLUDED.division,
    search_users   = EXCLUDED.search_users,
    search_count   = EXCLUDED.search_count,
    gmv            = EXCLUDED.gmv,
    add_cart_rate  = EXCLUDED.add_cart_rate,
    pay_rate       = EXCLUDED.pay_rate,
    no_result_rate = EXCLUDED.no_result_rate,
    source_file    = EXCLUDED.source_file,
    ingested_at    = now();
"""

SCHEMA = open(os.path.join(DIR, 'db_schema.sql'), encoding='utf-8').read()


def platform_from_name(path):
    """文件名含 APP（不分大小写）→ 'APP'，否则 '小程序'。"""
    return 'APP' if 'app' in os.path.basename(path).lower() else '小程序'


def main():
    # 解析 --platform 兜底默认值，其余视为文件路径
    args = sys.argv[1:]
    default_platform = None
    platform_explicit = False
    files_args = []
    force = False
    i = 0
    while i < len(args):
        if args[i] == '--platform':
            default_platform = args[i + 1]; platform_explicit = True; i += 2
        elif args[i].startswith('--platform='):
            default_platform = args[i].split('=', 1)[1]; platform_explicit = True; i += 1
        elif args[i] == '--force':
            force = True; i += 1
        else:
            files_args.append(args[i]); i += 1

    if files_args:
        files = [a if os.path.isabs(a) else os.path.join(DIR, a) for a in files_args]
    else:
        files = sorted(glob.glob(os.path.join(DIR, '*.xlsx')))
    if not files:
        sys.exit("未找到 xlsx 文件")

    conn = get_conn()
    cur = conn.cursor()
    cur.execute(SCHEMA)               # 确保表存在
    conn.commit()

    # 已入库文件名集合：默认跳过，避免重复导入（--force 可强制重导）
    cur.execute("SELECT DISTINCT source_file FROM ingest_log")
    done = {r[0] for r in cur.fetchall()}

    grand = 0
    for path in files:
        fname = os.path.basename(path)
        if fname in done and not force:
            print(f"– 跳过（已入库）：{fname}")
            continue
        # 显式 --platform 优先；否则按文件名自动判定（APP/小程序）
        plat = default_platform if platform_explicit else platform_from_name(path)
        rows, st = parse_xlsx(path, plat)
        if not rows:
            print(f"跳过（无数据）：{st['file']}")
            continue
        execute_values(cur, UPSERT, rows, page_size=5000)
        cur.execute(
            "INSERT INTO ingest_log (source_file, platform, row_count, date_min, date_max, keyword_cnt) "
            "VALUES (%s,%s,%s,%s,%s,%s)",
            (st['file'], st['platform'], st['n'], st['dmin'], st['dmax'], st['kw']),
        )
        conn.commit()
        grand += st['n']
        print(f"✓ {st['file']}: {st['n']} 行  [{st['platform']}]  {st['dmin']}~{st['dmax']}  {st['kw']} 词")

    cur.execute("SELECT count(*), min(date), max(date), count(distinct keyword) FROM keyword_daily")
    total, dmin, dmax, nkw = cur.fetchone()
    print(f"\n本次写入 {grand} 行。")
    print(f"数据库现有：{total} 行 · {dmin}~{dmax} · {nkw} 个关键词")
    cur.execute("SELECT platform, count(*) FROM keyword_daily GROUP BY platform ORDER BY platform")
    for plat, n in cur.fetchall():
        print(f"  - {plat}: {n} 行")
    cur.close(); conn.close()


if __name__ == '__main__':
    main()
